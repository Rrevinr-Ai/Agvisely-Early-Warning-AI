"""Excel-based crop advisory lookup (stage × weather → Bangla/English text)."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

from app.config import settings

_THRESHOLD_RE = re.compile(r"[<>]?\s*([\d.]+)")
_ASCII_NUM_RE = re.compile(r"[-+]?\d+(?:\.\d+)?")
_BN_DIGITS = str.maketrans("০১২৩৪৫৬৭৮৯", "0123456789")

_CROP_ALIASES = {
    "aman rice": "aman rice",
    "aman": "aman rice",
    "aman dhan": "aman rice",
    "আমন": "aman rice",
    "আমন ধান": "aman rice",
    "rice": "aman rice",
    "ধান": "aman rice",
}

# Farmer / speech variants → Excel English stage name
_STAGE_ALIASES = {
    "maximum tillering stage": "Maximum Tillering Stage",
    "maximum tillering": "Maximum Tillering Stage",
    "tillering": "Maximum Tillering Stage",
    "কুশি": "Maximum Tillering Stage",
    "কুশি পর্যায়": "Maximum Tillering Stage",
    "কুশি পর্যায়": "Maximum Tillering Stage",
    "কষি": "Maximum Tillering Stage",  # common STT/typo for কুশি
    "কষি পর্যায়": "Maximum Tillering Stage",
    "কষি পর্যায়ে": "Maximum Tillering Stage",
    "কষি পর্যায়ে": "Maximum Tillering Stage",
    "কষি পাওয়া": "Maximum Tillering Stage",  # STT: কষি পাওয়া যায় ≈ কুশি পর্যায়
    "কুষি": "Maximum Tillering Stage",  # STT variant
    "কুষি পর্যায়": "Maximum Tillering Stage",
    "কুষি পর্যায়ে": "Maximum Tillering Stage",
    "কুষি পর্যায়ে": "Maximum Tillering Stage",
    "পাওয়া যায় আছে": "Maximum Tillering Stage",  # with কষি nearby handled via কষি
    "booting stage": "Booting Stage",
    "booting": "Booting Stage",
    "থোড়": "Booting Stage",
    "থোড়": "Booting Stage",
    "থোড় পর্যায়": "Booting Stage",
    "থোড় পর্যায়": "Booting Stage",
    "থোড় পর্যায়ে": "Booting Stage",
}


@dataclass
class AdvisoryRow:
    crop_en: str
    crop_bn: str
    stage_en: str
    stage_bn: str
    start_date: Optional[date]
    end_date: Optional[date]
    stage_advisory: Optional[str]
    category: str
    dry_spell: bool
    min_threshold: Optional[float]
    max_threshold: Optional[float]
    min_exclusive: bool = False
    max_exclusive: bool = False
    advisory_en: str = ""
    advisory_bn: str = ""


@dataclass
class LoadedWorkbook:
    path: str
    rows: list[AdvisoryRow] = field(default_factory=list)


def _parse_threshold(value: Any) -> tuple[Optional[float], bool]:
    """Return (number, exclusive). exclusive=True when original had '>' or '<'."""
    if value is None or value == "":
        return None, False
    if isinstance(value, (int, float)):
        return float(value), False
    text = str(value).strip().translate(_BN_DIGITS)
    if not text:
        return None, False
    exclusive = text.startswith(">") or text.startswith("<")
    match = _THRESHOLD_RE.search(text.replace(",", ""))
    if not match:
        return None, False
    return float(match.group(1)), exclusive


def _parse_yes(value: Any) -> bool:
    if value is None:
        return False
    return str(value).strip().lower() in {"yes", "y", "true", "1", "হ্যাঁ"}


def _as_date(value: Any) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _normalize_crop(crop: str) -> Optional[str]:
    key = (crop or "").strip().lower()
    if not key:
        return None
    if key in _CROP_ALIASES:
        return _CROP_ALIASES[key]
    # fuzzy contains
    for alias, canonical in _CROP_ALIASES.items():
        if alias in key or key in alias:
            return canonical
    return None


def detect_crop_from_text(text: str) -> Optional[str]:
    """Infer crop from farmer speech (Bangla/English)."""
    if not text:
        return None
    lowered = text.lower()
    # Prefer specific Aman before generic rice
    markers = (
        ("আমন ধান", "aman rice"),
        ("আমন", "aman rice"),
        ("aman rice", "aman rice"),
        ("ধান", "aman rice"),
        ("rice", "aman rice"),
    )
    for marker, crop in markers:
        if marker in text or marker in lowered:
            return crop
    return None


def detect_stage_from_text(text: str) -> Optional[str]:
    """Infer growth stage from farmer speech, including common STT typos."""
    if not text:
        return None
    # Longer aliases first
    aliases = sorted(_STAGE_ALIASES.items(), key=lambda x: len(x[0]), reverse=True)
    lowered = text.lower()
    for alias, stage_en in aliases:
        if alias in text or alias.lower() in lowered:
            return stage_en
    return None


def normalize_stage_hint(stage: Optional[str]) -> Optional[str]:
    if not stage:
        return None
    key = stage.strip().lower()
    if key in _STAGE_ALIASES:
        return _STAGE_ALIASES[key]
    for alias, stage_en in _STAGE_ALIASES.items():
        if alias in key or key in alias:
            return stage_en
    # Allow exact Excel stage names
    return stage.strip()


def _month_day(d: date) -> tuple[int, int]:
    return d.month, d.day


def _in_season_window(today: date, start: Optional[date], end: Optional[date]) -> bool:
    """Match by month-day so 2021 Excel calendars work every year."""
    if not start or not end:
        return False
    t = _month_day(today)
    s = _month_day(start)
    e = _month_day(end)
    if s <= e:
        return s <= t <= e
    # wraps year-end (not used for Aman, but safe)
    return t >= s or t <= e


def _extract_numbers(text: str) -> list[float]:
    normalized = text.translate(_BN_DIGITS)
    return [float(x) for x in _ASCII_NUM_RE.findall(normalized)]


def extract_weather_metrics(weather: Optional[dict]) -> dict[str, Any]:
    """Best-effort numeric metrics from Agvisely or GPT weather payloads."""
    metrics: dict[str, Any] = {
        "temperature_c": None,
        "rainfall_mm": None,
        "is_dry_spell": False,
    }
    if not weather:
        return metrics

    for key in (
        "temperature_c",
        "temp_c",
        "temperature",
        "min_temperature",
        "max_temperature",
        "avg_temperature",
    ):
        value = weather.get(key)
        if isinstance(value, (int, float)):
            metrics["temperature_c"] = float(value)
            break
        if isinstance(value, str) and value.strip():
            nums = _extract_numbers(value)
            if nums:
                # Use midpoint of a range when present.
                metrics["temperature_c"] = sum(nums[:2]) / min(len(nums), 2)
                break

    for key in ("rainfall_mm", "rain_mm", "precipitation", "rainfall"):
        value = weather.get(key)
        if isinstance(value, (int, float)):
            metrics["rainfall_mm"] = float(value)
            break
        if isinstance(value, str) and value.strip():
            nums = _extract_numbers(value)
            if nums:
                metrics["rainfall_mm"] = nums[0]
                break

    outlook = " ".join(
        str(weather.get(k) or "")
        for k in ("rainfall_outlook", "weather_condition", "summary", "agent_speech")
    ).lower()
    dry_markers = ("dry spell", "no rainfall", "বৃষ্টিপাতের সম্ভাবনা নেই", "শুষ্ক", "<1")
    if any(marker in outlook for marker in dry_markers):
        metrics["is_dry_spell"] = True
        if metrics["rainfall_mm"] is None:
            metrics["rainfall_mm"] = 0.0

    return metrics


class ExcelAdvisoryService:
    def __init__(self) -> None:
        self._workbook: Optional[LoadedWorkbook] = None

    def _resolve_path(self) -> Optional[Path]:
        configured = Path(settings.EXCEL_ADVISORY_PATH)
        if configured.is_file():
            return configured
        # Fallbacks relative to project root
        root = Path(__file__).resolve().parents[2]
        candidates = [
            root / "Agvisely_Aman Rice 1.xlsx",
            root / "data" / "Agvisely_Aman Rice 1.xlsx",
        ]
        for path in candidates:
            if path.is_file():
                return path
        return None

    def reload(self) -> LoadedWorkbook:
        path = self._resolve_path()
        if not path:
            self._workbook = LoadedWorkbook(path="", rows=[])
            return self._workbook

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "openpyxl is required for Excel advisories. pip install openpyxl"
            ) from exc

        wb = load_workbook(path, data_only=True)
        ws = wb.active
        rows: list[AdvisoryRow] = []
        current = {
            "crop_en": "",
            "crop_bn": "",
            "stage_en": "",
            "stage_bn": "",
            "start_date": None,
            "end_date": None,
        }

        for values in ws.iter_rows(min_row=2, values_only=True):
            if not values or all(v is None or str(v).strip() == "" for v in values):
                continue

            (
                crop_en,
                crop_bn,
                stage_en,
                stage_bn,
                start_date,
                end_date,
                stage_advisory,
                category,
                dry_spell,
                min_raw,
                max_raw,
                advisory_en,
                advisory_bn,
            ) = (list(values) + [None] * 13)[:13]

            if crop_en:
                current["crop_en"] = str(crop_en).strip()
            if crop_bn:
                current["crop_bn"] = str(crop_bn).strip()
            if stage_en:
                current["stage_en"] = str(stage_en).strip()
            if stage_bn:
                current["stage_bn"] = str(stage_bn).strip()
            parsed_start = _as_date(start_date)
            parsed_end = _as_date(end_date)
            if parsed_start:
                current["start_date"] = parsed_start
            if parsed_end:
                current["end_date"] = parsed_end

            if not category:
                continue

            min_val, min_excl = _parse_threshold(min_raw)
            max_val, max_excl = _parse_threshold(max_raw)

            rows.append(
                AdvisoryRow(
                    crop_en=current["crop_en"],
                    crop_bn=current["crop_bn"],
                    stage_en=current["stage_en"],
                    stage_bn=current["stage_bn"],
                    start_date=current["start_date"],
                    end_date=current["end_date"],
                    stage_advisory=str(stage_advisory).strip().lower()
                    if stage_advisory
                    else None,
                    category=str(category).strip(),
                    dry_spell=_parse_yes(dry_spell),
                    min_threshold=min_val,
                    max_threshold=max_val,
                    min_exclusive=min_excl and str(min_raw).strip().startswith(">"),
                    max_exclusive=max_excl and str(max_raw).strip().startswith("<"),
                    advisory_en=(str(advisory_en).strip() if advisory_en else ""),
                    advisory_bn=(str(advisory_bn).strip() if advisory_bn else ""),
                )
            )

        self._workbook = LoadedWorkbook(path=str(path), rows=rows)
        return self._workbook

    def _ensure_loaded(self) -> LoadedWorkbook:
        if self._workbook is None:
            return self.reload()
        return self._workbook

    def available(self) -> bool:
        if not settings.EXCEL_ADVISORY_ENABLED:
            return False
        book = self._ensure_loaded()
        return bool(book.rows)

    def _crop_rows(self, crop: str) -> list[AdvisoryRow]:
        canonical = _normalize_crop(crop)
        if not canonical:
            return []
        book = self._ensure_loaded()
        return [
            row
            for row in book.rows
            if row.crop_en.strip().lower() == canonical
            or _normalize_crop(row.crop_en) == canonical
        ]

    def _stage_catalog(
        self, rows: list[AdvisoryRow]
    ) -> list[tuple[str, str, Optional[date], Optional[date]]]:
        seen: dict[str, tuple[str, str, Optional[date], Optional[date]]] = {}
        for row in rows:
            if not row.stage_en:
                continue
            key = row.stage_en.lower()
            if key not in seen:
                seen[key] = (row.stage_en, row.stage_bn, row.start_date, row.end_date)
        return list(seen.values())

    def _resolve_stage(
        self,
        rows: list[AdvisoryRow],
        today: date,
        stage_hint: Optional[str] = None,
    ) -> Optional[tuple[str, str, Optional[date], Optional[date]]]:
        catalog = self._stage_catalog(rows)
        if not catalog:
            return None

        hint = normalize_stage_hint(stage_hint)
        if hint:
            hint_l = hint.lower()
            for stage_en, stage_bn, start_date, end_date in catalog:
                if stage_en.lower() == hint_l or hint_l in stage_en.lower():
                    return stage_en, stage_bn, start_date, end_date
                if stage_bn and (hint in stage_bn or hint_l in stage_bn.lower()):
                    return stage_en, stage_bn, start_date, end_date

        # Calendar fallback when farmer did not state a stage
        active = []
        for stage_en, stage_bn, start_date, end_date in catalog:
            if _in_season_window(today, start_date, end_date):
                active.append((stage_en, stage_bn, start_date, end_date))
        if not active:
            return None
        active.sort(key=lambda s: s[2] or date.min, reverse=True)
        return active[0]

    def _matches_temperature(self, row: AdvisoryRow, temperature_c: Optional[float]) -> bool:
        if temperature_c is None:
            return False
        if row.min_threshold is not None and row.max_threshold is None:
            # cold advisory: below minimum
            return temperature_c < row.min_threshold
        if row.max_threshold is not None and row.min_threshold is None:
            # heat advisory: above maximum
            return temperature_c > row.max_threshold
        if row.min_threshold is not None and row.max_threshold is not None:
            return row.min_threshold <= temperature_c <= row.max_threshold
        return False

    def _matches_rainfall(self, row: AdvisoryRow, rainfall_mm: Optional[float], is_dry: bool) -> bool:
        if row.dry_spell:
            if rainfall_mm is not None:
                limit = row.min_threshold if row.min_threshold is not None else 1.0
                return rainfall_mm < limit
            return is_dry

        if rainfall_mm is None:
            return False

        low = row.min_threshold
        high = row.max_threshold

        # Very heavy band in Excel: only max filled (e.g. max=89 means >89 mm/d)
        if low is None and high is not None:
            return rainfall_mm > high

        if low is not None and row.min_exclusive:
            if not (rainfall_mm > low):
                return False
        elif low is not None and not (rainfall_mm >= low):
            return False

        if high is not None and not (rainfall_mm <= high):
            return False

        return low is not None or high is not None

    def _is_gap(self, row: AdvisoryRow) -> bool:
        return (
            row.stage_advisory == "yes"
            or row.category.lower() == "good agronomic practices"
        )

    def lookup(
        self,
        crop: str,
        weather: Optional[dict] = None,
        today: Optional[date] = None,
        district: Optional[str] = None,
        upazila: Optional[str] = None,
        stage: Optional[str] = None,
    ) -> Optional[dict]:
        """
        Return Excel advisory dict when crop+stage match; otherwise None (caller → GPT).

        If farmer states a stage (e.g. কুশি/কষি), that overrides the calendar window.
        """
        if not self.available():
            return None

        crop_rows = self._crop_rows(crop)
        if not crop_rows:
            return None

        today = today or date.today()
        resolved = self._resolve_stage(crop_rows, today, stage_hint=stage)
        if not resolved:
            return None

        stage_en, stage_bn, start_date, end_date = resolved
        stage_rows = [
            row
            for row in crop_rows
            if row.stage_en.lower() == stage_en.lower()
        ]

        metrics = extract_weather_metrics(weather)
        temperature_c = metrics.get("temperature_c")
        rainfall_mm = metrics.get("rainfall_mm")
        is_dry = bool(metrics.get("is_dry_spell"))

        # GPT/fallback weather numbers are estimates — do NOT fire Excel temp/rain thresholds from them.
        weather_source = str((weather or {}).get("source") or "").lower()
        has_numeric_obs = any(
            isinstance((weather or {}).get(k), (int, float))
            for k in (
                "temperature_c",
                "temp_c",
                "min_temperature",
                "max_temperature",
                "rainfall_mm",
                "rain_mm",
                "precipitation",
            )
        )
        trust_weather_thresholds = (
            weather_source not in {"gpt_backup", "fallback"} and has_numeric_obs
        )

        selected: list[AdvisoryRow] = []
        gap_rows: list[AdvisoryRow] = []
        weather_rows: list[AdvisoryRow] = []
        for row in stage_rows:
            if self._is_gap(row):
                gap_rows.append(row)
                selected.append(row)
                continue
            if not trust_weather_thresholds:
                continue
            category = row.category.lower()
            if category == "temperature" and self._matches_temperature(row, temperature_c):
                weather_rows.append(row)
                selected.append(row)
            elif category == "rainfall" and self._matches_rainfall(
                row, rainfall_mm, is_dry
            ):
                weather_rows.append(row)
                selected.append(row)

        # Always keep at least GAP if present; weather rows optional.
        if not selected:
            return None

        bangla_parts = [r.advisory_bn for r in selected if r.advisory_bn]
        english_parts = [r.advisory_en for r in selected if r.advisory_en]
        if not bangla_parts and not english_parts:
            return None

        gap_bn = [r.advisory_bn for r in gap_rows if r.advisory_bn]
        weather_bn = [r.advisory_bn for r in weather_rows if r.advisory_bn]
        agent_speech = "\n\n".join(bangla_parts) if bangla_parts else "\n\n".join(english_parts)
        book = self._ensure_loaded()

        return {
            "source": "excel",
            "crop": crop_rows[0].crop_en or crop,
            "crop_bn": crop_rows[0].crop_bn,
            "district": district,
            "upazila": upazila,
            "stage": stage_en,
            "stage_bn": stage_bn,
            "stage_source": "farmer" if stage else "calendar",
            "stage_start": start_date.isoformat() if start_date else None,
            "stage_end": end_date.isoformat() if end_date else None,
            "matched_categories": [r.category for r in selected],
            "weather_triggers": [r.category for r in weather_rows],
            "weather_metrics": metrics,
            "weather_thresholds_trusted": trust_weather_thresholds,
            "excel_file": book.path,
            "gap_advisory_bn": gap_bn,
            "weather_advisory_bn": weather_bn,
            "agent_speech": agent_speech,
            "message": agent_speech,
            "advisory_bn": bangla_parts,
            "advisory_en": english_parts,
            "disclaimer": "এই পরামর্শ Agvisely Excel নিয়ম থেকে নেওয়া হয়েছে।",
        }

    def bullet_id(self, text: str) -> str:
        import hashlib

        normalized = re.sub(r"\s+", " ", (text or "").strip())
        return hashlib.sha1(normalized.encode("utf-8")).hexdigest()[:12]

    def list_gap_bullets(self, hit: dict) -> list[dict]:
        gap = "\n".join(hit.get("gap_advisory_bn") or hit.get("advisory_bn") or [])
        weather_extra = "\n".join(hit.get("weather_advisory_bn") or [])
        bullets = self._extract_bullets(gap)
        if hit.get("weather_thresholds_trusted"):
            bullets.extend(self._extract_bullets(weather_extra))
        return [{"id": self.bullet_id(b), "text": b} for b in bullets]

    def select_bullets_for_dialog(
        self,
        hit: dict,
        *,
        intent: str = "general_crop",
        constraints: Optional[list[str]] = None,
        said_bullet_ids: Optional[list[str]] = None,
        max_bullets: int = 6,
    ) -> list[dict]:
        """Pick Excel bullets by intent/history/constraints (no LLM)."""
        constraints = constraints or []
        said = set(said_bullet_ids or [])
        all_bullets = self.list_gap_bullets(hit)

        pest_keys = (
            "পোকা",
            "রোগ",
            "বালাই",
            "মাজরা",
            "ফড়িং",
            "ফড়িং",
            "থ্রিপস",
            "ব্লাস্ট",
            "নাশক",
            "ফেরোমেন",
            "আলোক",
            "আইপিএম",
        )
        soil_keys = ("মাটি পরীক্ষা", "মাটি পরীক্ষ")
        alt_fert_keys = ("লিফ কালার", "এলসিসি", "LCC", "গুটি ইউরিয়া", "গুটি ইউরিয়া", "পরিমিত")

        def is_pest(b: dict) -> bool:
            return any(k in b["text"] for k in pest_keys)

        def is_soil_test(b: dict) -> bool:
            return any(k in b["text"] for k in soil_keys)

        def is_alt_fert(b: dict) -> bool:
            return any(k in b["text"] for k in alt_fert_keys)

        pool = list(all_bullets)

        if intent == "pest":
            focused = [b for b in pool if is_pest(b)]
            if focused:
                pool = focused
        elif intent == "constraint" or "no_money_for_soil_test" in constraints:
            # Drop soil-test insistence; prefer low-cost alternatives first
            pool = [b for b in pool if not is_soil_test(b)]
            alts = [b for b in all_bullets if is_alt_fert(b) and not is_soil_test(b)]
            rest = [b for b in pool if not is_alt_fert(b)]
            # Always surface cost-saving alternatives for this constraint, even if said before
            pool = alts + rest
        elif intent == "more_advice":
            # Prefer unseen bullets; skip already-said
            unseen = [b for b in pool if b["id"] not in said]
            if unseen:
                pool = unseen
            # Prefer later / IPM tips after fertilizer openers
            if len(pool) > 3:
                fert_openers = pool[:3]
                later = pool[3:]
                pool = later + fert_openers
        elif intent in {"weather", "weather_crop"}:
            # Only 1–2 field tips with weather
            pool = [b for b in pool if not is_soil_test(b)][:4]

        # Drop already-said when possible — except constraint alts which we force above
        if intent == "constraint" or "no_money_for_soil_test" in constraints:
            alts = [b for b in pool if is_alt_fert(b)]
            rest_fresh = [b for b in pool if not is_alt_fert(b) and b["id"] not in said]
            pool = alts + rest_fresh
        else:
            fresh = [b for b in pool if b["id"] not in said]
            if fresh:
                pool = fresh

        if "no_money_for_soil_test" in constraints:
            pool = [b for b in pool if not is_soil_test(b)]

        picked = pool[:max_bullets]
        if not picked and all_bullets:
            # Fallback: anything not soil-test if constrained
            fallback = all_bullets
            if "no_money_for_soil_test" in constraints:
                fallback = [b for b in fallback if not is_soil_test(b)] or all_bullets
            picked = [b for b in fallback if b["id"] not in said][:max_bullets] or fallback[:max_bullets]
        return picked

    def compose_spoken_advisory(
        self,
        hit: dict,
        user_message: str = "",
        max_bullets: int = 5,
        intent: str = "general_crop",
        constraints: Optional[list[str]] = None,
        said_bullet_ids: Optional[list[str]] = None,
    ) -> tuple[str, list[str]]:
        """Template fallback speech + list of used bullet ids."""
        crop = hit.get("crop_bn") or hit.get("crop") or "ধান"
        stage = hit.get("stage_bn") or hit.get("stage") or "বর্তমান পর্যায়"
        picked = self.select_bullets_for_dialog(
            hit,
            intent=intent,
            constraints=constraints,
            said_bullet_ids=said_bullet_ids,
            max_bullets=max_bullets,
        )
        ids = [b["id"] for b in picked]
        body = " ".join(b["text"] for b in picked)

        if intent == "constraint" or (constraints and "no_money_for_soil_test" in constraints):
            intro = (
                f"আপনার {crop} {stage}-এ আছে। মাটি পরীক্ষা ছাড়াই এখন যা করতে পারেন: "
            )
        elif intent == "pest":
            intro = f"{stage}-এ বালাই ব্যবস্থাপনার জন্য: "
        elif intent == "more_advice":
            intro = f"আরও কিছু কাজ {stage}-এ: "
        else:
            intro = f"আপনার {crop} এখন {stage}-এ আছে। "

        speech = (intro + body).strip()
        speech += " স্থানীয় কৃষি কর্মকর্তার পরামর্শ নিয়ে প্রয়োজনমতো ব্যবস্থা নিন।"
        return speech, ids

    @staticmethod
    def _extract_bullets(text: str) -> list[str]:
        if not text:
            return []
        parts: list[str] = []
        for raw in re.split(r"[\n•]+", text):
            line = re.sub(r"\s+", " ", raw).strip(" -–—\t")
            if len(line) < 12:
                continue
            # Drop section headers that are too short / labels only
            if line.endswith(":") and len(line) < 40:
                continue
            parts.append(line if line.endswith("।") else line + "।")
        return parts


excel_advisory_service = ExcelAdvisoryService()
